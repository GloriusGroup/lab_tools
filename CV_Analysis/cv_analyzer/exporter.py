"""Export CV data and analysis results to Excel."""

from pathlib import Path
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    Workbook = None


def _auto_width(ws):
    """Set each column width to fit its longest value."""
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 3
        ws.column_dimensions[col[0].column_letter].width = max_len


_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                           fill_type="solid")
_SECTION_FONT = Font(bold=True, size=11)


def _write_header(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    return cell


def export_to_excel(potential, current, result, metadata,
                    output_path, ref_label="Ag/AgCl"):
    """Write two-sheet Excel workbook.

    Sheet 1 – "Raw Data": potential and current columns.
    Sheet 2 – "Analysis": extracted peak values, reversibility, and
              measurement metadata.

    Parameters
    ----------
    potential, current : array-like
        CV data arrays (already in output reference).
    result : CVResult
        Analysis result object.
    metadata : dict
        From ``parser.extract_metadata``.
    output_path : str or Path
        Destination .xlsx path.
    ref_label : str
        Output reference electrode label.
    """
    if Workbook is None:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = Workbook()

    # ===================================================================
    # Sheet 1: Raw Data
    # ===================================================================
    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    _write_header(ws_raw, 1, 1, f"Potential (V vs {ref_label})")
    _write_header(ws_raw, 1, 2, f"Current ({metadata.get('current_unit', 'uA')})")

    for i, (v, c) in enumerate(zip(potential, current), start=2):
        ws_raw.cell(row=i, column=1, value=float(v))
        ws_raw.cell(row=i, column=2, value=float(c))

    _auto_width(ws_raw)

    # ===================================================================
    # Sheet 2: Analysis
    # ===================================================================
    ws_ana = wb.create_sheet("Analysis")
    row = 1

    # --- Section: Measurement Metadata ---
    ws_ana.cell(row=row, column=1, value="Measurement Metadata").font = _SECTION_FONT
    row += 1

    meta_rows = [
        ("Identifier", metadata.get("identifier", "")),
        ("Source file", metadata.get("filename", "")),
        ("Data points", metadata.get("n_points")),
        ("Potential unit", metadata.get("potential_unit", "V")),
        ("Current unit", metadata.get("current_unit", "uA")),
        ("Start potential (V)", metadata.get("start_potential")),
        ("End potential (V)", metadata.get("end_potential")),
        ("Min potential (V)", metadata.get("min_potential")),
        ("Max potential (V)", metadata.get("max_potential")),
        ("Potential step (V)", metadata.get("potential_step_V")),
        ("Vertex potentials (V)",
         ", ".join(f"{v:.4f}" for v in metadata.get("vertex_potentials", []))),
        ("Number of sweeps", metadata.get("n_sweeps")),
        ("Output reference electrode", ref_label),
    ]
    for label, value in meta_rows:
        _write_header(ws_ana, row, 1, label)
        ws_ana.cell(row=row, column=2, value=value)
        row += 1

    row += 1  # blank spacer

    # --- Section: Reversibility ---
    ws_ana.cell(row=row, column=1,
                value="Reversibility Assessment").font = _SECTION_FONT
    row += 1
    _write_header(ws_ana, row, 1, "Classification")
    ws_ana.cell(row=row, column=2,
                value="Reversible (quasi-reversible)" if result.is_reversible
                else "Irreversible")
    row += 1

    if result.is_reversible and result.standard_potential is not None:
        _write_header(ws_ana, row, 1, f"Standard potential E0 (V vs {ref_label})")
        ws_ana.cell(row=row, column=2, value=round(result.standard_potential, 4))
        row += 1
        _write_header(ws_ana, row, 1, "Peak separation dEp (mV)")
        ws_ana.cell(row=row, column=2,
                    value=round(result.peak_separation * 1000, 1))
        row += 1

    row += 1  # blank spacer

    # --- Section: Detected Peaks ---
    ws_ana.cell(row=row, column=1, value="Detected Peaks").font = _SECTION_FONT
    row += 1

    if not result.peaks:
        ws_ana.cell(row=row, column=1, value="No peaks detected.")
        row += 1
    else:
        # Table header
        headers = [
            "#", "Type",
            f"Onset (V vs {ref_label})",
            f"Ep (V vs {ref_label})",
            f"Ip ({metadata.get('current_unit', 'uA')})",
            f"Ep/2 (V vs {ref_label})",
            "|Ep - Ep/2| (mV)",
        ]
        for c, h in enumerate(headers, 1):
            _write_header(ws_ana, row, c, h)
        row += 1

        ox_count = 0
        red_count = 0
        for pk in result.peaks:
            if pk.peak_type == "anodic":
                ox_count += 1
                tag = f"Ox{ox_count}"
                ptype = "Oxidation"
            else:
                red_count += 1
                tag = f"Red{red_count}"
                ptype = "Reduction"

            delta = abs(pk.potential - pk.half_peak_potential) * 1000
            values = [
                tag, ptype,
                round(pk.onset_potential, 4),
                round(pk.potential, 4),
                round(pk.current, 4),
                round(pk.half_peak_potential, 4),
                round(delta, 1),
            ]
            for c, v in enumerate(values, 1):
                ws_ana.cell(row=row, column=c, value=v)
            row += 1

    _auto_width(ws_ana)

    # Save
    wb.save(str(output_path))
    return Path(output_path)
